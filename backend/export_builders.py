"""Build Markdown and Excel exports for failures-questions and referrals-by-channel."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Shared styles ────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SECTION_FONT = Font(bold=True, size=10, color="1F3864")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _auto_width(ws, min_w: int = 14, max_w: int = 55):
    for col in ws.columns:
        mx = min_w
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx


# ── Helpers ──────────────────────────────────────────────────────────────────

def _build_fecha_map(df: pd.DataFrame) -> dict[str, str]:
    """Build a thread_id -> fecha (YYYY-MM-DD) lookup from the main df.

    Uses the first row per thread which is fast and avoids per-thread queries.
    """
    if df is None or df.empty or "fecha" not in df.columns:
        return {}
    first = df.drop_duplicates(subset="thread_id", keep="first")[["thread_id", "fecha"]]
    result: dict[str, str] = {}
    for _, row in first.iterrows():
        val = row["fecha"]
        result[row["thread_id"]] = str(val)[:10] if pd.notna(val) else ""
    return result


# ── Dimension filtering helper ───────────────────────────────────────────────

def _filter_by_dimension(
    frame: pd.DataFrame,
    dimension: str,
    value: str,
    subcats: Optional[set] = None,
    product_threads: Optional[set] = None,
) -> pd.DataFrame:
    """Filter a failures or referrals DataFrame by dimension+value."""
    if frame is None or frame.empty:
        return frame

    if dimension == "product":
        if product_threads is not None:
            return frame[frame["thread_id"].isin(product_threads)]
        return frame.head(0)

    # dimension == "category": value is a macro name
    if subcats:
        return frame[frame["intencion"].isin(subcats)]
    return frame[frame["intencion"] == value]


# ── Markdown builder ─────────────────────────────────────────────────────────

def build_failures_questions_md(
    failures_df: pd.DataFrame,
    df: pd.DataFrame,
    referrals_df: pd.DataFrame,
    dimension: str,
    value: str,
    period_start: Optional[str] = None,
    period_end: Optional[str] = None,
) -> str:
    """Build a Markdown report of user questions that caused bot failures.

    Only includes failures with criteria 'Respuesta de incapacidad del bot'
    (i.e. the bot explicitly said it cannot help). Includes thread_id, the
    user question, the bot's failure response, whether it was redirected, and
    to which channel.
    """
    dim_label = "Producto" if dimension == "product" else "Categoria"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Only keep real failures (bot incapacity), not sentiment or repetition
    # Then filter out noise: greetings, surveys, very short messages
    if failures_df is not None and not failures_df.empty:
        incap_df = failures_df[
            failures_df["criteria"].str.contains("incapacidad", case=False, na=False)
        ].copy()
        # Remove rows where the user question is pure noise (survey, very short,
        # or a greeting with no real content — max ~25 chars for a greeting)
        if not incap_df.empty:
            noise = incap_df["last_user_message"].str.strip().str.lower()
            is_noise = (
                noise.str.contains(r"^\[survey\]", case=False, na=False)
                | (noise.str.len() < 6)
                | (
                    noise.str.match(
                        r"^(hola|hey|hi|ok|gracias|si|no|dale|listo"
                        r"|buenos?\s*d[ií]as?"
                        r"|buenas?\s*(tardes?|noches?|dias?)?"
                        r"|muy\s+buenas?\s*(tardes?|d[ií]as?)?"
                        r"|hola\s+(buenas?|buenos?)\s*(tardes?|d[ií]as?|noches?)?"
                        r"|muchas\s+gracias"
                        r"|mil\s+gracias"
                        r"|ola\s+buenas?\s*(tardes?)?"
                        r")\s*[.!,?]*$",
                        na=False,
                    )
                )
            )
            incap_df = incap_df[~is_noise]
    else:
        incap_df = pd.DataFrame()

    total = len(incap_df)

    # Build referral lookup: thread_id -> channel
    ref_lookup: dict[str, str] = {}
    if referrals_df is not None and not referrals_df.empty:
        for _, r in referrals_df.iterrows():
            ch = str(r.get("channel", ""))
            label = {"digital": "Digital", "serviline": "Servilinea", "office": "Oficina"}.get(ch, ch)
            ref_lookup[r["thread_id"]] = label

    lines: list[str] = []
    lines.append("# Reporte de Preguntas sin Informacion del Asistente")
    lines.append("")
    lines.append(f"**{dim_label}:** {value}")
    if period_start or period_end:
        lines.append(f"**Periodo:** {period_start or '...'} a {period_end or '...'}")
    lines.append(f"**Generado:** {now}")
    lines.append(f"**Total de preguntas donde el asistente no tenia informacion:** {total}")
    lines.append("")

    if total == 0:
        lines.append("No se encontraron preguntas sin informacion para esta seleccion.")
        return "\n".join(lines)

    # Resolve fecha from main df (vectorized lookup)
    fecha_map = _build_fecha_map(df)

    # Add resolved fecha to incap_df for sorting
    incap_df["_fecha"] = incap_df["thread_id"].map(fecha_map).fillna("")

    # Group by subcategory
    lines.append("## Detalle por subcategoria")
    lines.append("")

    grouped = incap_df.groupby("intencion", sort=False)
    sorted_groups = sorted(grouped, key=lambda g: len(g[1]), reverse=True)

    for cat_name, group in sorted_groups:
        # Sort by fecha within each group
        group = group.sort_values("_fecha")

        lines.append(f"### {cat_name} ({len(group)} preguntas)")
        lines.append("")
        lines.append("| Thread ID | Fecha | Pregunta del Usuario | Respuesta del Bot | Redirigido |")
        lines.append("|-----------|-------|---------------------|-------------------|------------|")

        for _, row in group.iterrows():
            tid = str(row.get("thread_id", ""))
            fecha = row.get("_fecha", "")
            user_q = str(row.get("last_user_message", "")).replace("|", " ").replace("\n", " ")[:150]
            bot_r = str(row.get("last_ai_message", "")).replace("|", " ").replace("\n", " ")[:150]
            redir = ref_lookup.get(tid, "No")
            tid_short = tid[-12:] if len(tid) > 12 else tid
            lines.append(f"| {tid_short} | {fecha} | {user_q} | {bot_r} | {redir} |")

        lines.append("")

    lines.append("---")
    lines.append("*Reporte generado automaticamente desde el dashboard del asistente.*")
    return "\n".join(lines)


# ── Excel builder ────────────────────────────────────────────────────────────

def _write_conversations_sheet(
    ws,
    thread_ids: list[str],
    df: pd.DataFrame,
    meta_df: pd.DataFrame,
    referrals_df: pd.DataFrame,
    meta_col_intencion: str = "intencion",
):
    """Write a sheet with full conversations for the given thread_ids.

    Each thread becomes a group of rows (one per message), separated by a
    section header row with the thread metadata.
    ``ws`` must be an already-created worksheet.
    """

    headers = [
        "Thread ID",
        "Fecha",
        "Subcategoria",
        "Quien",
        "Mensaje",
        "Sentimiento",
        "Redirigido",
        "Canal",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    # Build referral lookup
    ref_lookup: dict[str, str] = {}
    if referrals_df is not None and not referrals_df.empty:
        for _, r in referrals_df.iterrows():
            ch = str(r.get("channel", ""))
            ref_lookup[r["thread_id"]] = {"digital": "Digital", "serviline": "Servilinea", "office": "Oficina"}.get(ch, ch)

    # Build meta lookup
    meta_lookup: dict[str, dict] = {}
    if meta_df is not None and not meta_df.empty:
        for _, r in meta_df.iterrows():
            meta_lookup[r["thread_id"]] = {
                "intencion": str(r.get(meta_col_intencion, "")),
                "sentiment": str(r.get("sentiment", "")),
            }

    if not thread_ids:
        ws.cell(row=2, column=1, value="Sin datos para esta seleccion")
        _auto_width(ws)
        return

    # Build fecha map and sort thread_ids by date
    fecha_map = _build_fecha_map(df)
    thread_ids_sorted = sorted(thread_ids, key=lambda t: fecha_map.get(t, ""))

    # Pre-filter df to only relevant threads and sort chronologically
    relevant_df = df[df["thread_id"].isin(set(thread_ids))]
    if "rowid" in relevant_df.columns:
        relevant_df = relevant_df.sort_values("rowid")
    grouped_msgs = {tid: grp for tid, grp in relevant_df.groupby("thread_id")}

    row = 2
    for tid in thread_ids_sorted:
        thread = grouped_msgs.get(tid)
        if thread is None or thread.empty:
            continue

        fecha = fecha_map.get(tid, "")
        meta = meta_lookup.get(tid, {})
        intencion = meta.get("intencion", "")
        sentiment = meta.get("sentiment", "")
        canal = ref_lookup.get(tid, "")
        redirigido = "Si" if canal else "No"

        # Section header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cell = ws.cell(row=row, column=1, value=f"{tid} | {fecha} | {intencion} | Redirigido: {redirigido} {canal}")
        cell.font = _SECTION_FONT
        cell.fill = _SECTION_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = _SECTION_FILL
        row += 1

        # Message rows in chronological order (only human and ai)
        for _, m in thread.iterrows():
            msg_type = m.get("type", "")
            if msg_type not in ("human", "ai"):
                continue
            who = "Usuario" if msg_type == "human" else "Bot"
            text = str(m.get("text", ""))

            ws.cell(row=row, column=1, value=tid[-12:])
            ws.cell(row=row, column=2, value=fecha)
            ws.cell(row=row, column=3, value=intencion)
            ws.cell(row=row, column=4, value=who)
            ws.cell(row=row, column=5, value=text)
            ws.cell(row=row, column=6, value=sentiment)
            ws.cell(row=row, column=7, value=redirigido)
            ws.cell(row=row, column=8, value=canal)

            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = _THIN_BORDER
                ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    _auto_width(ws)


def build_failures_referrals_excel(
    failures_df: pd.DataFrame,
    referrals_df: pd.DataFrame,
    df: pd.DataFrame,
    dimension: str,
    value: str,
) -> bytes:
    """Build an Excel workbook with full conversations.

    Sheet 1 - Sin Informacion: full conversations for threads where the bot lacked information.
    Sheet 2 - Canal Digital: full conversations for threads redirected to digital.
    Sheet 3 - Canal Servilinea: full conversations for threads redirected to servilinea.
    Sheet 4 - Canal Oficina: full conversations for threads redirected to office.
    """
    wb = Workbook()

    # ── Sheet 1: Fallos (only incapacity) ────────────────────────────────
    if failures_df is not None and not failures_df.empty:
        incap_df = failures_df[
            failures_df["criteria"].str.contains("incapacidad", case=False, na=False)
        ]
        fail_tids = list(incap_df["thread_id"].unique())
    else:
        incap_df = pd.DataFrame()
        fail_tids = []

    ws_fail = wb.active
    ws_fail.title = "Sin Informacion"
    _write_conversations_sheet(ws_fail, fail_tids, df, incap_df, referrals_df)

    # ── Sheets 2-4: By channel ───────────────────────────────────────────
    channel_map = [
        ("Canal Digital", "digital"),
        ("Canal Servilinea", "serviline"),
        ("Canal Oficina", "office"),
    ]

    for sheet_name, channel_value in channel_map:
        if referrals_df is not None and not referrals_df.empty:
            channel_df = referrals_df[referrals_df["channel"] == channel_value]
            ch_tids = list(channel_df["thread_id"].unique())
        else:
            channel_df = pd.DataFrame()
            ch_tids = []

        ws = wb.create_sheet(sheet_name)
        _write_conversations_sheet(ws, ch_tids, df, channel_df, referrals_df)

    # ── Save to bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
