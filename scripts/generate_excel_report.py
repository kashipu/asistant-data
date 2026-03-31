"""Generate Excel methodology annex for the monthly report."""
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from backend.engine import DataEngine
from backend.dashboard_metrics import get_extended_funnel
from backend.metrics import get_general_kpis
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

e = DataEngine()
df = e.df
refs = e.get_referrals()
fails = e.get_failures()
ref_set = set(refs.thread_id.unique())
fail_set = set(fails.thread_id.unique())

funnel = get_extended_funnel(df)
kpis = get_general_kpis(df)

wb = Workbook()

# ── Styles ──
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
section_font = Font(bold=True, size=11, color="1F3864")
number_fmt = "#,##0"
pct_fmt = "0.0%"
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def style_section(ws, row, cols, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = section_font
    cell.fill = section_fill
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = section_fill
        ws.cell(row=row, column=c).border = thin_border


def auto_width(ws, min_w=12, max_w=60):
    for col in ws.columns:
        mx = min_w
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx


# ════════════════════════════════════════════════════════
# HOJA 1: Metodologia del Embudo
# ════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Embudo - Metodologia"

headers = [
    "ID Metrica",
    "Nombre",
    "Valor",
    "Porcentaje",
    "Base de Calculo",
    "Formula / Como se Calcula",
    "Fuente de Datos",
    "Interpretacion",
]
for c, h in enumerate(headers, 1):
    ws1.cell(row=1, column=c, value=h)
style_header(ws1, 1, len(headers))

formulas = {
    "total": "COUNT(DISTINCT thread_id) en tabla messages",
    "greeting_only": "Convs donde TODOS los msgs humanos tienen <= 5 palabras. Filtra saludos, ok, gracias",
    "active": "Convs con > 2 msgs humanos, excluyendo las de solo saludo. = total - greeting_only - convs con <= 2 msgs humanos",
    "self_service": "Activas que NO aparecen en tabla referrals. = activas - redirigidas",
    "referred": "Activas que SI aparecen en tabla referrals (deteccion por keywords: servilinea, oficina, digital)",
    "surveyed": 'Convs donde existe un msg con texto que contiene "[survey]"',
    "answered": "De las encuestadas, las que tienen respuesta Me fue util o No me fue util",
    "useful": 'Msgs humanos con texto "[survey] Me fue util la informacion"',
    "not_useful": 'Msgs humanos con texto "[survey] No me fue util la informacion"',
    "failures": "Activas detectadas por 3 criterios: (1) bot dice no puedo/no tengo/no es posible, (2) usuario repite misma pregunta, (3) >50% msgs con sentimiento negativo",
    "waste": "Interseccion: thread_id en (no_util) AND thread_id en (referrals)",
    "referred_failed": "Interseccion: thread_id en (failures) AND thread_id en (referrals)",
    "with_product": 'Activas donde product_yaml IS NOT NULL y != ""',
    "general": 'Activas donde product_yaml IS NULL o = ""',
    "arrived_seeking_advisor": 'Convs donde categoria_yaml del PRIMER msg = "Escalamiento a Asesor"',
    "organic_escalation": "Convs en referrals MENOS las que llegaron buscando asesor",
}

metrics = funnel["metrics"]
row = 2
for m in metrics:
    ws1.cell(row=row, column=1, value=m["id"])
    ws1.cell(row=row, column=2, value=m["label"])
    ws1.cell(row=row, column=3, value=m["count"])
    ws1.cell(row=row, column=3).number_format = number_fmt
    ws1.cell(row=row, column=4, value=m["pct"] / 100)
    ws1.cell(row=row, column=4).number_format = pct_fmt
    ws1.cell(row=row, column=5, value=m.get("base_label", ""))
    ws1.cell(row=row, column=6, value=formulas.get(m["id"], m.get("explanation", "")))
    ws1.cell(row=row, column=7, value="dashboard_metrics.get_extended_funnel()")
    ws1.cell(row=row, column=8, value=m.get("explanation", ""))

    for c in range(1, len(headers) + 1):
        ws1.cell(row=row, column=c).border = thin_border
        ws1.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    # Breakdown if exists
    if "breakdown" in m:
        for bd in m["breakdown"]:
            row += 1
            ws1.cell(row=row, column=2, value=f"  -> {bd['label']}")
            ws1.cell(row=row, column=3, value=bd["count"])
            ws1.cell(row=row, column=3).number_format = number_fmt
            ws1.cell(
                row=row,
                column=6,
                value="Subclasificacion por keywords en referral_response",
            )
            for c in range(1, len(headers) + 1):
                ws1.cell(row=row, column=c).border = thin_border

    # Correlation if exists
    if "correlation" in m:
        cor = m["correlation"]
        row += 1
        ws1.cell(row=row, column=2, value=f"  [corr] {cor['label']}")
        ws1.cell(row=row, column=3, value=cor["count"])
        ws1.cell(row=row, column=3).number_format = number_fmt
        ws1.cell(row=row, column=4, value=cor["pct"] / 100)
        ws1.cell(row=row, column=4).number_format = pct_fmt
        for c in range(1, len(headers) + 1):
            ws1.cell(row=row, column=c).border = thin_border

    row += 1

auto_width(ws1)


# ════════════════════════════════════════════════════════
# HOJA 2: KPIs Operacionales
# ════════════════════════════════════════════════════════
ws2 = wb.create_sheet("KPIs Operacionales")

headers2 = ["KPI", "Valor", "Unidad", "Formula", "Fuente", "Interpretacion"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, len(headers2))

kpi_rows = [
    ("Total Conversaciones", kpis["total_conversations"], "convs", "COUNT(DISTINCT thread_id)", "messages", "Total bruto de hilos de chat"),
    ("Total Mensajes", kpis["total_messages"], "msgs", "COUNT(*) en messages", "messages", "Incluye human, ai, tool"),
    ("Mensajes Humanos", kpis["messages_by_type"]["human"], "msgs", "COUNT(*) WHERE type='human'", "messages", "Lo que escribio el cliente"),
    ("Mensajes IA", kpis["messages_by_type"]["ai"], "msgs", "COUNT(*) WHERE type='ai'", "messages", "Respuestas generadas por el bot"),
    ("Mensajes Tool", kpis["messages_by_type"]["tool"], "msgs", "COUNT(*) WHERE type='tool'", "messages", "Llamadas internas a herramientas"),
    ("Usuarios Unicos", kpis["total_users"], "users", "COUNT(DISTINCT client_ip)", "messages", "IPs unicas como proxy de usuarios"),
    ("Promedio msgs/conv", kpis["avg_messages_per_thread"], "msgs", "total_messages / total_conversations", "calculado", "Profundidad promedio de interaccion"),
    ("Promedio msgs humanos/conv", kpis["avg_human_messages_per_thread"], "msgs", "SUM(human_msgs) / total_conversations", "calculado", "Cuanto escribe el cliente en promedio"),
    ("Mediana msgs/conv", kpis["median_messages_per_thread"], "msgs", "MEDIAN(msgs por thread)", "calculado", "Valor central, menos afectado por outliers"),
    ("Tasa de Abandono", kpis["abandonment_rate"], "%", "(convs con <=1 msg humano) / total x 100", "calculado", "% que se va sin interactuar"),
    ("Promedio convs/usuario", kpis["avg_conversations_per_user"], "convs", "total_conversations / total_users", "calculado", "Recurrencia del usuario"),
    ("Tokens Entrada Total", kpis["total_input_tokens"], "tokens", "SUM(input_tokens)", "messages", "Costo de contexto enviado al modelo"),
    ("Tokens Salida Total", kpis["total_output_tokens"], "tokens", "SUM(output_tokens)", "messages", "Costo de respuestas generadas"),
    ("Tokens Entrada/msg IA", kpis["avg_input_tokens_per_ai_msg"], "tokens", "total_input / COUNT(ai msgs)", "calculado", "Contexto promedio por respuesta"),
    ("Tokens Salida/msg IA", kpis["avg_output_tokens_per_ai_msg"], "tokens", "total_output / COUNT(ai msgs)", "calculado", "Longitud promedio de respuesta"),
]

for i, (name, val, unit, formula, source, interp) in enumerate(kpi_rows, 2):
    ws2.cell(row=i, column=1, value=name)
    ws2.cell(row=i, column=2, value=val)
    ws2.cell(row=i, column=2).number_format = number_fmt if isinstance(val, int) else "0.00"
    ws2.cell(row=i, column=3, value=unit)
    ws2.cell(row=i, column=4, value=formula)
    ws2.cell(row=i, column=5, value=source)
    ws2.cell(row=i, column=6, value=interp)
    for c in range(1, len(headers2) + 1):
        ws2.cell(row=i, column=c).border = thin_border

auto_width(ws2)


# ════════════════════════════════════════════════════════
# HOJA 3: Categorias Detalladas
# ════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Categorias Detalladas")

headers3 = [
    "Macrocategoria",
    "Subcategoria",
    "Conversaciones",
    "% del Total",
    "Derivadas",
    "% Derivacion",
    "Fallos",
    "% Fallos",
    "% Negativo",
    "Ejemplo Pregunta 1",
    "Ejemplo Pregunta 2",
    "Ejemplo Pregunta 3",
    "Casuistica / Que se encuentra",
]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, len(headers3))

total_threads = df.thread_id.nunique()
macros_all = (
    df[df.macro_yaml.notna() & (df.macro_yaml != "")]
    .groupby("macro_yaml")
    .thread_id.nunique()
    .sort_values(ascending=False)
)

row = 2
for macro_name, macro_count in macros_all.items():
    style_section(
        ws3,
        row,
        len(headers3),
        f"{macro_name} ({macro_count:,} convs - {macro_count/total_threads*100:.1f}%)",
    )
    row += 1

    macro_df = df[df.macro_yaml == macro_name]
    subs = (
        macro_df.groupby("categoria_yaml")
        .thread_id.nunique()
        .sort_values(ascending=False)
    )

    for sub_name, sub_count in subs.items():
        sub_df = df[df.categoria_yaml == sub_name]
        sub_threads_set = set(
            macro_df[macro_df.categoria_yaml == sub_name].thread_id.unique()
        )
        sub_ref = len(sub_threads_set & ref_set)
        sub_fail = len(sub_threads_set & fail_set)

        sub_human = sub_df[sub_df.type == "human"]
        neg = sub_human[sub_human.sentiment == "negativo"].shape[0]
        total_s = max(len(sub_human), 1)

        # Examples
        first_msgs = sub_df[sub_df.type == "human"].drop_duplicates(
            subset="thread_id", keep="first"
        )
        first_msgs = first_msgs[first_msgs.text.str.len() > 10]
        first_msgs = first_msgs[
            ~first_msgs.text.str.contains(
                r"survey|hola |buenos |buenas ", case=False, na=False
            )
        ]
        top_msgs = first_msgs.text.str.strip().value_counts().head(3)
        examples = list(top_msgs.index)

        # Casuistica
        casui = []
        if sub_ref / max(sub_count, 1) > 0.85:
            casui.append("Alta derivacion")
        if sub_fail / max(sub_count, 1) > 0.35:
            casui.append("Alto fallo del bot")
        if neg / total_s > 0.25:
            casui.append("Alta frustracion")
        if sub_count > 1000:
            casui.append("Alto volumen")

        ws3.cell(row=row, column=1, value=macro_name)
        ws3.cell(row=row, column=2, value=sub_name)
        ws3.cell(row=row, column=3, value=sub_count)
        ws3.cell(row=row, column=3).number_format = number_fmt
        ws3.cell(row=row, column=4, value=sub_count / total_threads)
        ws3.cell(row=row, column=4).number_format = pct_fmt
        ws3.cell(row=row, column=5, value=sub_ref)
        ws3.cell(row=row, column=5).number_format = number_fmt
        ws3.cell(row=row, column=6, value=sub_ref / max(sub_count, 1))
        ws3.cell(row=row, column=6).number_format = pct_fmt
        ws3.cell(row=row, column=7, value=sub_fail)
        ws3.cell(row=row, column=7).number_format = number_fmt
        ws3.cell(row=row, column=8, value=sub_fail / max(sub_count, 1))
        ws3.cell(row=row, column=8).number_format = pct_fmt
        ws3.cell(row=row, column=9, value=neg / total_s)
        ws3.cell(row=row, column=9).number_format = pct_fmt
        ws3.cell(row=row, column=10, value=examples[0] if len(examples) > 0 else "")
        ws3.cell(row=row, column=11, value=examples[1] if len(examples) > 1 else "")
        ws3.cell(row=row, column=12, value=examples[2] if len(examples) > 2 else "")
        ws3.cell(
            row=row,
            column=13,
            value="; ".join(casui) if casui else "Normal",
        )

        for c in range(1, len(headers3) + 1):
            ws3.cell(row=row, column=c).border = thin_border
            ws3.cell(row=row, column=c).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        row += 1

auto_width(ws3, max_w=45)


# ════════════════════════════════════════════════════════
# HOJA 4: Productos Detallados
# ════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Productos Detallados")

headers4 = [
    "Producto",
    "Categoria Cruzada",
    "Conversaciones",
    "% del Producto",
    "Derivadas",
    "% Derivacion",
    "Fallos",
    "% Fallos",
    "Ejemplo Pregunta 1",
    "Ejemplo Pregunta 2",
    "Ejemplo Pregunta 3",
]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, 1, len(headers4))

products_all = (
    df[df.product_yaml.notna() & (df.product_yaml != "")]
    .groupby("product_yaml")
    .thread_id.nunique()
    .sort_values(ascending=False)
)

row = 2
for prod_name, prod_count in products_all.items():
    style_section(
        ws4,
        row,
        len(headers4),
        f"{prod_name} ({prod_count:,} convs - {prod_count/total_threads*100:.1f}%)",
    )
    row += 1

    prod_df = df[df.product_yaml == prod_name]
    cats = (
        prod_df[prod_df.categoria_yaml.notna()]
        .groupby("categoria_yaml")
        .thread_id.nunique()
        .sort_values(ascending=False)
    )

    for cat_name, cat_count in cats.head(10).items():
        cat_in_prod = prod_df[prod_df.categoria_yaml == cat_name]
        cat_threads = set(cat_in_prod.thread_id.unique())
        cat_ref = len(cat_threads & ref_set)
        cat_fail = len(cat_threads & fail_set)

        first_msgs = cat_in_prod[cat_in_prod.type == "human"].drop_duplicates(
            subset="thread_id", keep="first"
        )
        first_msgs = first_msgs[first_msgs.text.str.len() > 10]
        first_msgs = first_msgs[
            ~first_msgs.text.str.contains(
                r"survey|hola |buenos |buenas ", case=False, na=False
            )
        ]
        top = list(first_msgs.text.str.strip().value_counts().head(3).index)

        ws4.cell(row=row, column=1, value=prod_name)
        ws4.cell(row=row, column=2, value=cat_name)
        ws4.cell(row=row, column=3, value=cat_count)
        ws4.cell(row=row, column=3).number_format = number_fmt
        ws4.cell(row=row, column=4, value=cat_count / max(prod_count, 1))
        ws4.cell(row=row, column=4).number_format = pct_fmt
        ws4.cell(row=row, column=5, value=cat_ref)
        ws4.cell(row=row, column=5).number_format = number_fmt
        ws4.cell(row=row, column=6, value=cat_ref / max(cat_count, 1))
        ws4.cell(row=row, column=6).number_format = pct_fmt
        ws4.cell(row=row, column=7, value=cat_fail)
        ws4.cell(row=row, column=7).number_format = number_fmt
        ws4.cell(row=row, column=8, value=cat_fail / max(cat_count, 1))
        ws4.cell(row=row, column=8).number_format = pct_fmt
        ws4.cell(row=row, column=9, value=top[0] if len(top) > 0 else "")
        ws4.cell(row=row, column=10, value=top[1] if len(top) > 1 else "")
        ws4.cell(row=row, column=11, value=top[2] if len(top) > 2 else "")

        for c in range(1, len(headers4) + 1):
            ws4.cell(row=row, column=c).border = thin_border
            ws4.cell(row=row, column=c).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        row += 1

auto_width(ws4, max_w=45)


# ════════════════════════════════════════════════════════
# HOJA 5: Gasto de Valor
# ════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Gasto de Valor")

headers5 = [
    "Categoria",
    "Conversaciones Gasto",
    "% del Gasto Total",
    "Formula",
    "Interpretacion",
]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=c, value=h)
style_header(ws5, 1, len(headers5))

waste_data = funnel["waste_by_category"]
for i, w in enumerate(waste_data, 2):
    ws5.cell(row=i, column=1, value=w["category"])
    ws5.cell(row=i, column=2, value=w["count"])
    ws5.cell(row=i, column=2).number_format = number_fmt
    ws5.cell(row=i, column=3, value=w["pct_of_waste"] / 100)
    ws5.cell(row=i, column=3).number_format = pct_fmt
    ws5.cell(
        row=i,
        column=4,
        value='thread_id IN (no_util) AND thread_id IN (referrals) AND categoria_yaml = "'
        + w["category"]
        + '"',
    )
    ws5.cell(
        row=i,
        column=5,
        value="Doble fallo: el bot no resolvio Y escalo al usuario. Valor perdido.",
    )
    for c in range(1, len(headers5) + 1):
        ws5.cell(row=i, column=c).border = thin_border

auto_width(ws5, max_w=55)


# ════════════════════════════════════════════════════════
# HOJA 6: Temporal
# ════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Temporal")

ws6.cell(row=1, column=1, value="Hora")
ws6.cell(row=1, column=2, value="Mensajes Humanos")
ws6.cell(row=1, column=3, value="% del Total")
ws6.cell(row=1, column=4, value="Formula")
style_header(ws6, 1, 4)

hourly = df[df.type == "human"].groupby("hora").size().sort_index()
for i, (h, c) in enumerate(hourly.items(), 2):
    ws6.cell(row=i, column=1, value=f"{int(h):02d}:00")
    ws6.cell(row=i, column=2, value=c)
    ws6.cell(row=i, column=2).number_format = number_fmt
    ws6.cell(row=i, column=3, value=c / hourly.sum())
    ws6.cell(row=i, column=3).number_format = pct_fmt
    ws6.cell(
        row=i,
        column=4,
        value="COUNT(*) WHERE type='human' AND hora=" + str(int(h)),
    )
    for cc in range(1, 5):
        ws6.cell(row=i, column=cc).border = thin_border

# Day of week
r = len(hourly) + 4
ws6.cell(row=r, column=1, value="Dia")
ws6.cell(row=r, column=2, value="Mensajes Humanos")
ws6.cell(row=r, column=3, value="% del Total")
style_header(ws6, r, 3)

df_temp = df.copy()
df_temp["dow"] = pd.to_datetime(df_temp.fecha).dt.day_name()
dow = df_temp[df_temp.type == "human"].groupby("dow").size()
dow_order = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
]
for d in dow_order:
    if d in dow.index:
        r += 1
        ws6.cell(row=r, column=1, value=d)
        ws6.cell(row=r, column=2, value=dow[d])
        ws6.cell(row=r, column=2).number_format = number_fmt
        ws6.cell(row=r, column=3, value=dow[d] / dow.sum())
        ws6.cell(row=r, column=3).number_format = pct_fmt
        for cc in range(1, 4):
            ws6.cell(row=r, column=cc).border = thin_border

auto_width(ws6)


# ════════════════════════════════════════════════════════
# HOJA 7: Fallos Detallados
# ════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Fallos Detallados")

headers7 = [
    "Categoria de Fallo",
    "Convs con Fallo",
    "% del Total Fallos",
    "Criterio Principal",
    "Formula de Deteccion",
]
for c, h in enumerate(headers7, 1):
    ws7.cell(row=1, column=c, value=h)
style_header(ws7, 1, len(headers7))

if "intencion" in fails.columns:
    fail_cat = fails.groupby("intencion").size().sort_values(ascending=False)
    total_fails = fail_cat.sum()
    for i, (name, count) in enumerate(fail_cat.items(), 2):
        ws7.cell(row=i, column=1, value=name)
        ws7.cell(row=i, column=2, value=count)
        ws7.cell(row=i, column=2).number_format = number_fmt
        ws7.cell(row=i, column=3, value=count / total_fails)
        ws7.cell(row=i, column=3).number_format = pct_fmt

        cat_fails = fails[fails.intencion == name]
        criteria_counts = (
            cat_fails.criteria.str.split(",").explode().str.strip().value_counts()
        )
        primary = criteria_counts.index[0] if len(criteria_counts) > 0 else ""
        ws7.cell(row=i, column=4, value=primary)
        ws7.cell(
            row=i,
            column=5,
            value="detect_failures(df) -> criteria matches por keywords en last_ai_message + repeticion + sentiment",
        )

        for c in range(1, len(headers7) + 1):
            ws7.cell(row=i, column=c).border = thin_border

auto_width(ws7)


# ════════════════════════════════════════════════════════
# HOJA 8: Derivaciones por Canal
# ════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Derivaciones por Canal")

headers8 = [
    "Canal",
    "Conversaciones",
    "% del Total Derivaciones",
    "Metodo de Deteccion",
    "Keywords de Clasificacion",
]
for c, h in enumerate(headers8, 1):
    ws8.cell(row=1, column=c, value=h)
style_header(ws8, 1, len(headers8))

channel_data = refs.channel.value_counts()
total_refs = channel_data.sum()
keywords_map = {
    "digital": "app, banca movil, banca virtual, web, digital, internet",
    "serviline": "servilinea, linea telefonica, llamar, numero telefonico, contacto telefonico",
    "office": "oficina, sucursal, punto de atencion, sede, presencialmente",
}
for i, (ch, cnt) in enumerate(channel_data.items(), 2):
    ws8.cell(row=i, column=1, value=ch)
    ws8.cell(row=i, column=2, value=cnt)
    ws8.cell(row=i, column=2).number_format = number_fmt
    ws8.cell(row=i, column=3, value=cnt / total_refs)
    ws8.cell(row=i, column=3).number_format = pct_fmt
    ws8.cell(
        row=i,
        column=4,
        value="detect_referrals(df) -> clasifica por keywords en referral_response del bot",
    )
    ws8.cell(row=i, column=5, value=keywords_map.get(ch, ""))
    for c in range(1, len(headers8) + 1):
        ws8.cell(row=i, column=c).border = thin_border

auto_width(ws8)

# ════════════════════════════════════════════════════════
# HOJA 9: Sentimiento por Categoria
# ════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Sentimiento por Categoria")

headers9 = [
    "Macrocategoria",
    "Msgs Positivo",
    "Msgs Neutral",
    "Msgs Negativo",
    "Total Msgs",
    "% Negativo",
    "Formula",
]
for c, h in enumerate(headers9, 1):
    ws9.cell(row=1, column=c, value=h)
style_header(ws9, 1, len(headers9))

human_df = df[
    (df.type == "human") & (df.macro_yaml.notna()) & (df.macro_yaml != "")
]
sent_macro = (
    human_df.groupby(["macro_yaml", "sentiment"]).size().unstack(fill_value=0)
)
sent_macro["total"] = sent_macro.sum(axis=1)
sent_macro = sent_macro.sort_values("total", ascending=False)

for i, (name, row_data) in enumerate(sent_macro.iterrows(), 2):
    ws9.cell(row=i, column=1, value=name)
    ws9.cell(row=i, column=2, value=int(row_data.get("positivo", 0)))
    ws9.cell(row=i, column=2).number_format = number_fmt
    ws9.cell(row=i, column=3, value=int(row_data.get("neutral", 0)))
    ws9.cell(row=i, column=3).number_format = number_fmt
    ws9.cell(row=i, column=4, value=int(row_data.get("negativo", 0)))
    ws9.cell(row=i, column=4).number_format = number_fmt
    ws9.cell(row=i, column=5, value=int(row_data["total"]))
    ws9.cell(row=i, column=5).number_format = number_fmt
    neg_pct = row_data.get("negativo", 0) / max(row_data["total"], 1)
    ws9.cell(row=i, column=6, value=neg_pct)
    ws9.cell(row=i, column=6).number_format = pct_fmt
    ws9.cell(
        row=i,
        column=7,
        value="COUNT(*) WHERE type='human' GROUP BY macro_yaml, sentiment",
    )
    for c in range(1, len(headers9) + 1):
        ws9.cell(row=i, column=c).border = thin_border

auto_width(ws9)


# Save
output_path = "e:/code/asistente-tablero/anexo_metodologico_marzo_2026.xlsx"
wb.save(output_path)
print(f"Excel saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print("Done!")
