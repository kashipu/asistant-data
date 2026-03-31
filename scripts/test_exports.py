"""Comprehensive tests for the failures/referrals export builders."""
import sys
import io
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from backend.engine import DataEngine
from backend.export_builders import (
    build_failures_questions_md,
    build_failures_referrals_excel,
    _filter_by_dimension,
)
import pandas as pd
from openpyxl import load_workbook

e = DataEngine()
df = e.df
failures_df = e.get_failures()
referrals_df = e.get_referrals()

errors = []
warnings = []

# ===================================================================
# TEST 1: _filter_by_dimension — product
# ===================================================================
print("TEST 1: Filter by product (Tarjeta de Credito)")
prod_threads = set(
    df[df["product_yaml"] == "Tarjeta de Crédito"]["thread_id"].unique()
)
f_prod = _filter_by_dimension(
    failures_df, "product", "Tarjeta de Crédito", product_threads=prod_threads
)
r_prod = _filter_by_dimension(
    referrals_df, "product", "Tarjeta de Crédito", product_threads=prod_threads
)

if f_prod.empty:
    errors.append("T1: Product filter returned 0 failures for TDC")
else:
    print(f"  Failures: {len(f_prod)}, Referrals: {len(r_prod)}")
    bad = set(f_prod.thread_id) - prod_threads
    if bad:
        errors.append(f"T1: {len(bad)} failure threads not in product set")
    bad_r = set(r_prod.thread_id) - prod_threads
    if bad_r:
        errors.append(f"T1: {len(bad_r)} referral threads not in product set")
    print("  All thread_ids belong to product: OK")

# ===================================================================
# TEST 2: _filter_by_dimension — category
# ===================================================================
print("\nTEST 2: Filter by category (Tarjetas)")
subcats = set(
    df[df.macro_yaml == "Tarjetas"]["categoria_yaml"].dropna().unique()
)
f_cat = _filter_by_dimension(failures_df, "category", "Tarjetas", subcats)
r_cat = _filter_by_dimension(referrals_df, "category", "Tarjetas", subcats)

if f_cat.empty:
    errors.append("T2: Category filter returned 0 failures for Tarjetas")
else:
    print(f"  Failures: {len(f_cat)}, Referrals: {len(r_cat)}")
    bad_int = set(f_cat.intencion) - subcats
    if bad_int:
        errors.append(f"T2: Failure intenciones outside subcats: {bad_int}")
    print(f"  Subcats: {subcats}")
    print("  All intenciones in subcats: OK")

# ===================================================================
# TEST 3: _filter_by_dimension — empty / None inputs
# ===================================================================
print("\nTEST 3: Filter with empty/None inputs")
empty_df = pd.DataFrame(columns=failures_df.columns)
result = _filter_by_dimension(empty_df, "product", "X", product_threads=set())
if not result.empty:
    errors.append("T3: Expected empty result for empty input")
else:
    print("  Empty input -> empty output: OK")

result2 = _filter_by_dimension(None, "product", "X", product_threads=set())
if result2 is not None:
    errors.append("T3: Expected None for None input")
else:
    print("  None input -> None output: OK")

# ===================================================================
# TEST 4: Markdown — content quality (product)
# ===================================================================
print("\nTEST 4: Markdown report quality (product: TDC)")
md = build_failures_questions_md(
    f_prod, df, r_prod, "product", "Tarjeta de Crédito", "2026-01-28", "2026-03-03"
)
lines = md.split("\n")

if "Reporte de Preguntas sin Informacion" not in lines[0]:
    errors.append("T4: Missing title")
if "Tarjeta de Crédito" not in md:
    errors.append("T4: Missing product name in MD")
if "2026-01-28" not in md:
    errors.append("T4: Missing period start")
print("  Header: OK")

# No None dates
if "| None " in md:
    errors.append("T4: Found 'None' dates in markdown")
else:
    print("  No None dates: OK")

# No pure greetings
data_rows = [
    l
    for l in lines
    if l.startswith("| ") and "Thread" not in l and "---" not in l
]
greeting_pat = re.compile(
    r"\| (hola|buenas tardes|buenos dias|buenas noches) \|", re.IGNORECASE
)
greeting_rows = [l for l in data_rows if greeting_pat.search(l)]
if greeting_rows:
    warnings.append(f"T4: {len(greeting_rows)} pure greetings in MD")
else:
    print(f"  No pure greetings: OK ({len(data_rows)} data rows)")

# No [survey]
survey_rows = [l for l in data_rows if "[survey]" in l.lower()]
if survey_rows:
    errors.append(f"T4: {len(survey_rows)} survey messages in MD")
else:
    print("  No survey messages: OK")

# Thread ID column
if "Thread ID" not in md:
    errors.append("T4: Missing Thread ID column")
else:
    print("  Thread ID column: OK")

# Redirigido column with channel labels
if "Redirigido" not in md:
    errors.append("T4: Missing Redirigido column")
else:
    has_channel = any(
        "Digital" in l or "Servilinea" in l or "Oficina" in l for l in data_rows
    )
    if not has_channel:
        warnings.append("T4: No channel labels found in Redirigido column")
    else:
        print("  Redirigido column with channels: OK")

# ===================================================================
# TEST 5: Markdown — category filter
# ===================================================================
print("\nTEST 5: Markdown for category (Pagos)")
subcats_pagos = set(
    df[df.macro_yaml == "Pagos"]["categoria_yaml"].dropna().unique()
)
f_pagos = _filter_by_dimension(failures_df, "category", "Pagos", subcats_pagos)
r_pagos = _filter_by_dimension(referrals_df, "category", "Pagos", subcats_pagos)
md_pagos = build_failures_questions_md(
    f_pagos, df, r_pagos, "category", "Pagos"
)
if "Categoria" not in md_pagos:
    errors.append("T5: Missing 'Categoria' label")
if "Pagos" not in md_pagos:
    errors.append("T5: Missing 'Pagos' in content")
pagos_data = [
    l
    for l in md_pagos.split("\n")
    if l.startswith("| ") and "Thread" not in l and "---" not in l
]
print(f"  Pagos fallos data rows: {len(pagos_data)}")
if len(pagos_data) == 0:
    warnings.append("T5: No failure data rows for Pagos")
else:
    print("  OK")

# ===================================================================
# TEST 6: Markdown — empty case
# ===================================================================
print("\nTEST 6: Markdown with no failures")
empty_f = pd.DataFrame(columns=failures_df.columns)
md_empty = build_failures_questions_md(
    empty_f, df, r_prod, "product", "Producto X"
)
if "No se encontraron preguntas sin informacion" not in md_empty:
    errors.append("T6: Missing empty-state message")
else:
    print("  Empty failures -> proper message: OK")

# ===================================================================
# TEST 7: Excel — structure (category Tarjetas)
# ===================================================================
print("\nTEST 7: Excel report structure (category: Tarjetas)")
xlsx_bytes = build_failures_referrals_excel(
    f_cat, r_cat, df, "category", "Tarjetas"
)

wb = load_workbook(io.BytesIO(xlsx_bytes))
expected_sheets = ["Sin Informacion", "Canal Digital", "Canal Servilinea", "Canal Oficina"]
if wb.sheetnames != expected_sheets:
    errors.append(f"T7: Wrong sheets: {wb.sheetnames}")
else:
    print(f"  4 sheets: {wb.sheetnames} OK")

expected_headers = [
    "Thread ID",
    "Fecha",
    "Subcategoria",
    "Quien",
    "Mensaje",
    "Sentimiento",
    "Redirigido",
    "Canal",
]
for sn in expected_sheets:
    ws = wb[sn]
    actual = [ws.cell(row=1, column=c).value for c in range(1, 9)]
    if actual != expected_headers:
        errors.append(f"T7: {sn} headers: {actual}")
print("  All headers correct: OK")

# ===================================================================
# TEST 8: Excel — full conversations in Fallos
# ===================================================================
print("\nTEST 8: Excel Sin Informacion sheet — full conversations")
ws_fail = wb["Sin Informacion"]
row_count = ws_fail.max_row
if row_count <= 1:
    errors.append("T8: Sin Informacion sheet empty")
else:
    print(f"  Sin Informacion rows: {row_count - 1}")

# Check both Usuario and Bot messages exist
who_values = set()
for r_idx in range(2, min(row_count + 1, 200)):
    val = ws_fail.cell(row=r_idx, column=4).value
    if val and val in ("Usuario", "Bot"):
        who_values.add(val)
if "Usuario" not in who_values:
    errors.append("T8: No Usuario messages in Fallos")
if "Bot" not in who_values:
    errors.append("T8: No Bot messages in Fallos")
if who_values == {"Usuario", "Bot"}:
    print("  Full conversations (Usuario + Bot): OK")

# Check Fecha is not None
fecha_set = set()
for r_idx in range(2, min(row_count + 1, 50)):
    val = ws_fail.cell(row=r_idx, column=2).value
    if val and str(val).strip() and str(val) != "None":
        fecha_set.add(str(val))
if not fecha_set:
    errors.append("T8: No valid fechas in Excel")
else:
    has_none = any(
        str(ws_fail.cell(row=r, column=2).value) == "None"
        for r in range(2, min(row_count + 1, 50))
        if ws_fail.cell(row=r, column=2).value
    )
    if has_none:
        errors.append("T8: Found 'None' fechas in Excel")
    else:
        print(f"  Fechas valid (sample: {list(fecha_set)[:2]}): OK")

# ===================================================================
# TEST 9: Excel — channel sheets have data and correct channel
# ===================================================================
print("\nTEST 9: Excel channel sheets")
channel_checks = {
    "Canal Digital": "Digital",
    "Canal Servilinea": "Servilinea",
    "Canal Oficina": "Oficina",
}
for sn, expected_chan in channel_checks.items():
    ws = wb[sn]
    rows = ws.max_row
    print(f"  {sn}: {rows - 1} rows")
    if rows <= 1:
        warnings.append(f"T9: {sn} is empty")
        continue
    # Check Redirigido=Si and Canal matches
    found_match = False
    for r_idx in range(2, min(rows + 1, 50)):
        redir = ws.cell(row=r_idx, column=7).value
        canal = ws.cell(row=r_idx, column=8).value
        if redir == "Si" and canal == expected_chan:
            found_match = True
            break
    if not found_match:
        errors.append(
            f"T9: {sn} has no rows with Redirigido=Si, Canal={expected_chan}"
        )
    else:
        print(f"    Channel label '{expected_chan}': OK")

# ===================================================================
# TEST 10: Excel — empty referrals
# ===================================================================
print("\nTEST 10: Excel with empty referrals")
empty_r = pd.DataFrame(columns=referrals_df.columns)
xlsx_empty = build_failures_referrals_excel(
    f_cat.head(2), empty_r, df, "category", "Tarjetas"
)
wb2 = load_workbook(io.BytesIO(xlsx_empty))
ws2_dig = wb2["Canal Digital"]
cell_val = ws2_dig.cell(row=2, column=1).value
if "Sin datos" not in str(cell_val or ""):
    errors.append(f"T10: Empty channel sheet says '{cell_val}' not 'Sin datos'")
else:
    print("  Empty channels -> 'Sin datos': OK")

# ===================================================================
# TEST 11: Excel — product filter
# ===================================================================
print("\nTEST 11: Excel for product (Credito de Libre Inversion)")
prod2 = set(
    df[df["product_yaml"] == "Crédito de Libre Inversión"]["thread_id"].unique()
)
f2 = _filter_by_dimension(failures_df, "product", "CLI", product_threads=prod2)
r2 = _filter_by_dimension(referrals_df, "product", "CLI", product_threads=prod2)
xlsx2 = build_failures_referrals_excel(
    f2, r2, df, "product", "Crédito de Libre Inversión"
)
wb3 = load_workbook(io.BytesIO(xlsx2))
print(f"  Sin Informacion rows: {wb3['Sin Informacion'].max_row - 1}")
print(f"  Canal Digital rows: {wb3['Canal Digital'].max_row - 1}")
if wb3["Sin Informacion"].max_row <= 1 and len(f2) > 0:
    errors.append("T11: Sin Informacion sheet empty despite having failures")
else:
    print("  OK")

# ===================================================================
# TEST 12: Backend endpoints registered
# ===================================================================
print("\nTEST 12: Backend endpoints")
from backend.main import app

routes = [r.path for r in app.routes if hasattr(r, "path")]
for ep in [
    "/api/reports/export/failures-questions-markdown",
    "/api/reports/export/failures-referrals-excel",
]:
    if ep not in routes:
        errors.append(f"T12: Missing endpoint {ep}")
print("  Both endpoints registered: OK")

# ===================================================================
# TEST 13: Frontend TypeScript
# ===================================================================
print("\nTEST 13: Frontend compilation (skipped - run 'npx tsc --noEmit' manually)")

# ===================================================================
# RESULTS
# ===================================================================
print("\n" + "=" * 60)
if errors:
    print(f"ERRORS: {len(errors)}")
    for em in errors:
        print(f"  [FAIL] {em}")
else:
    print("ALL TESTS PASSED")

if warnings:
    print(f"\nWARNINGS: {len(warnings)}")
    for w in warnings:
        print(f"  [WARN] {w}")

sys.exit(1 if errors else 0)
